from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, g, has_request_context
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import io
import os
import threading
import time
import shutil
import webbrowser
import base64
from xml.sax.saxutils import escape

# Ruta base absoluta del proyecto (evita problemas con cwd relativo)
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
DB_PATH = os.path.join(DATA_DIR, 'app.db')
UPLOADS_DIR = os.path.join(BASE_DIR, 'static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{DB_PATH}"
app.config['UPLOAD_FOLDER'] = UPLOADS_DIR
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-please-change')

# Inicializar la extensión sin forzar conexión hasta que el app context exista
db = SQLAlchemy(app)


def parse_currency(value):
    if value is None:
        return 0.0
    digits = ''.join(ch for ch in str(value) if ch.isdigit())
    return float(digits) if digits else 0.0


def classify_income_method(method_value):
    text = (method_value or '').strip().lower()
    if 'efectivo' in text:
        return 'efectivo'
    if 'nequi' in text:
        return 'nequi'
    account_keywords = ['cuenta', 'transfer', 'daviplata', 'bancolombia', 'banco', 'qr']
    if any(keyword in text for keyword in account_keywords):
        return 'cuentas'
    return 'cuentas'


def get_brand_name():
    cfg = get_current_config()
    if cfg and cfg.company_name:
        return cfg.company_name
    return 'CellSite'


def get_payment_accounts():
    cfg = get_current_config()
    raw = (cfg.payment_accounts or '') if cfg else ''
    accounts = [line.strip() for line in raw.splitlines() if line.strip()]
    return accounts


def get_current_config():
    """Cachea la configuración en el ciclo de request para evitar consultas repetidas."""
    if has_request_context():
        if hasattr(g, '_cached_config'):
            return g._cached_config
        g._cached_config = Config.query.first()
        return g._cached_config
    return Config.query.first()


@app.context_processor
def inject_branding():
    current_user = get_current_user()
    current_perms = ROLE_PERMISSIONS.get(current_user.role, set()) if current_user else set()
    return {
        'brand_name': get_brand_name(),
        'payment_accounts': get_payment_accounts(),
        'user_perms': current_perms,
        'user_role': current_user.role if current_user else None,
        'role_labels': ROLE_LABELS,
    }


class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(30), nullable=False, default='vendedor')
    active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, server_default=db.func.now())


class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    id_nit = db.Column(db.String(80))
    phone = db.Column(db.String(50))
    email = db.Column(db.String(200))
    address = db.Column(db.String(300))
    notes = db.Column(db.Text)
    sales = db.relationship('Sale', backref='client', cascade='all, delete-orphan', order_by='Sale.date')


class InventoryContainer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    description = db.Column(db.String(300))
    active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, server_default=db.func.now())


class Inventory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    device_type = db.Column(db.String(60), nullable=False)
    imei = db.Column(db.String(100), unique=True, nullable=False)
    imei_secondary = db.Column(db.String(100))
    model = db.Column(db.String(200))
    color = db.Column(db.String(50))
    invoice_number = db.Column(db.String(100))
    purchase_price = db.Column(db.Numeric(10,2), nullable=False, default=0)
    price = db.Column(db.Numeric(10,2), nullable=False, default=0)
    container_id = db.Column(db.Integer, db.ForeignKey('inventory_container.id'))
    status = db.Column(db.String(30), nullable=False, default='disponible')
    created_at = db.Column(db.DateTime, server_default=db.func.now())
    sold_at = db.Column(db.DateTime)
    container = db.relationship('InventoryContainer', backref='items')


class InventoryMovement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    inventory_id = db.Column(db.Integer, db.ForeignKey('inventory.id'), nullable=False)
    action = db.Column(db.String(30), nullable=False)  # ingreso|traslado|venta|devolucion
    from_container_id = db.Column(db.Integer, db.ForeignKey('inventory_container.id'))
    to_container_id = db.Column(db.Integer, db.ForeignKey('inventory_container.id'))
    note = db.Column(db.String(300))
    username = db.Column(db.String(80))
    created_at = db.Column(db.DateTime, server_default=db.func.now())

    inventory = db.relationship('Inventory', backref='movements')
    from_container = db.relationship('InventoryContainer', foreign_keys=[from_container_id])
    to_container = db.relationship('InventoryContainer', foreign_keys=[to_container_id])


class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sale_id = db.Column(db.Integer, db.ForeignKey('sale.id'), nullable=False)
    inventory_id = db.Column(db.Integer, db.ForeignKey('inventory.id'))
    imei = db.Column(db.String(100), unique=True, nullable=False)
    imei_secondary = db.Column(db.String(100))
    device_type = db.Column(db.String(60))
    color = db.Column(db.String(50))
    invoice_number = db.Column(db.String(100))
    model = db.Column(db.String(200))
    purchase_price = db.Column(db.Numeric(10,2), default=0)
    price = db.Column(db.Numeric(10,2))
    paid = db.Column(db.Boolean, default=False)  # Marca si el producto está completamente pagado
    inventory = db.relationship('Inventory', backref='products')


class Sale(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'), nullable=False)
    sale_price = db.Column(db.Numeric(10,2), nullable=False)
    date = db.Column(db.DateTime, server_default=db.func.now())
    initial_amount = db.Column(db.Numeric(10,2), nullable=False)
    balance = db.Column(db.Numeric(10,2), nullable=False)
    status = db.Column(db.String(30), default='activo')
    products = db.relationship('Product', backref='sale', cascade='all, delete-orphan')
    payments = db.relationship('Payment', backref='sale', cascade='all, delete-orphan')


class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sale_id = db.Column(db.Integer, db.ForeignKey('sale.id'), nullable=False)
    date = db.Column(db.DateTime, server_default=db.func.now())
    amount = db.Column(db.Numeric(10,2), nullable=False)
    method = db.Column(db.String(100))
    account_from = db.Column(db.String(200))
    account_to = db.Column(db.String(200))
    note = db.Column(db.String(300))
    username = db.Column(db.String(80))


class Return(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sale_id = db.Column(db.Integer, db.ForeignKey('sale.id'), nullable=False)
    date = db.Column(db.DateTime, server_default=db.func.now())
    reason = db.Column(db.String(300))
    refund_amount = db.Column(db.Numeric(10,2), default=0)

# añadir relación desde Sale a Return
Sale.returns = db.relationship('Return', backref='sale', cascade='all, delete-orphan')


class Config(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(200))
    company_phone = db.Column(db.String(50))
    company_email = db.Column(db.String(200))
    company_address = db.Column(db.String(300))
    invoice_description = db.Column(db.Text)
    invoice_template = db.Column(db.Text)
    receipt_template = db.Column(db.Text)
    payment_accounts = db.Column(db.Text)
    logo_path = db.Column(db.String(300))  # ruta relativa a la imagen del logo


class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, server_default=db.func.now())
    username = db.Column(db.String(80))
    role = db.Column(db.String(30))
    action = db.Column(db.String(100), nullable=False)
    resource_type = db.Column(db.String(50))
    resource_id = db.Column(db.Integer)
    detail = db.Column(db.Text)
    ip = db.Column(db.String(45))


# ---- Permisos por rol ----
ROLE_PERMISSIONS = {
    'admin': {
        'ver_clientes', 'crear_clientes',
        'ver_ventas', 'crear_ventas', 'ver_pagos', 'registrar_pagos',
        'ver_inventario', 'ver_precio_compra', 'agregar_inventario', 'mover_inventario',
        'ver_estadisticas', 'exportar_excel',
        'ver_configuracion', 'gestionar_usuarios', 'ver_auditoria',
    },
    'vendedor': {
        'ver_clientes', 'crear_clientes',
        'ver_ventas', 'crear_ventas', 'ver_pagos', 'registrar_pagos',
        'ver_inventario',
    },
    'bodeguero': {
        'ver_inventario', 'ver_precio_compra', 'agregar_inventario', 'mover_inventario',
    },
    'readonly': {
        'ver_clientes', 'ver_ventas', 'ver_pagos',
        'ver_inventario', 'ver_precio_compra', 'ver_estadisticas',
    },
}

ROLE_LABELS = {
    'admin': 'Administrador',
    'vendedor': 'Vendedor',
    'bodeguero': 'Bodeguero',
    'readonly': 'Solo lectura',
}


def get_current_user():
    uname = session.get('user')
    if not uname:
        return None
    if has_request_context():
        if hasattr(g, '_cached_user'):
            return g._cached_user
        g._cached_user = User.query.filter_by(username=uname).first()
        return g._cached_user
    return User.query.filter_by(username=uname).first()


def get_user_perms():
    user = get_current_user()
    if not user:
        return set()
    return ROLE_PERMISSIONS.get(user.role, set())


def has_perm(perm):
    return perm in get_user_perms()


def require_perm(perm):
    """Decorator: redirect to index with flash if user lacks permission."""
    from functools import wraps
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not session.get('user'):
                return redirect(url_for('login'))
            if not has_perm(perm):
                flash('No tienes permiso para acceder a esta sección.', 'error')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return wrapped
    return decorator


def log_action(action, resource_type=None, resource_id=None, detail=None):
    user = get_current_user()
    try:
        ip = request.remote_addr
    except RuntimeError:
        ip = None
    entry = AuditLog(
        username=user.username if user else session.get('user'),
        role=user.role if user else None,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        detail=detail,
        ip=ip,
    )
    db.session.add(entry)
    # commit is handled by caller


def ensure_schema_compatibility():
    product_columns = {
        row[1]
        for row in db.session.execute(db.text('PRAGMA table_info(product)')).fetchall()
    }
    if 'inventory_id' not in product_columns:
        db.session.execute(db.text('ALTER TABLE product ADD COLUMN inventory_id INTEGER'))
    if 'device_type' not in product_columns:
        db.session.execute(db.text('ALTER TABLE product ADD COLUMN device_type VARCHAR(60)'))
    if 'purchase_price' not in product_columns:
        db.session.execute(db.text('ALTER TABLE product ADD COLUMN purchase_price NUMERIC(10,2) DEFAULT 0'))
    if 'imei_secondary' not in product_columns:
        db.session.execute(db.text('ALTER TABLE product ADD COLUMN imei_secondary VARCHAR(100)'))

    payment_columns = {
        row[1]
        for row in db.session.execute(db.text('PRAGMA table_info(payment)')).fetchall()
    }
    if 'account_from' not in payment_columns:
        db.session.execute(db.text('ALTER TABLE payment ADD COLUMN account_from VARCHAR(200)'))
    if 'account_to' not in payment_columns:
        db.session.execute(db.text('ALTER TABLE payment ADD COLUMN account_to VARCHAR(200)'))
    if 'username' not in payment_columns:
        db.session.execute(db.text('ALTER TABLE payment ADD COLUMN username VARCHAR(80)'))

    inventory_columns = {
        row[1]
        for row in db.session.execute(db.text('PRAGMA table_info(inventory)')).fetchall()
    }
    if 'purchase_price' not in inventory_columns:
        db.session.execute(db.text('ALTER TABLE inventory ADD COLUMN purchase_price NUMERIC(10,2) DEFAULT 0'))
    if 'imei_secondary' not in inventory_columns:
        db.session.execute(db.text('ALTER TABLE inventory ADD COLUMN imei_secondary VARCHAR(100)'))
    if 'container_id' not in inventory_columns:
        db.session.execute(db.text('ALTER TABLE inventory ADD COLUMN container_id INTEGER'))

    client_columns = {
        row[1]
        for row in db.session.execute(db.text('PRAGMA table_info(client)')).fetchall()
    }
    if 'id_nit' not in client_columns:
        db.session.execute(db.text('ALTER TABLE client ADD COLUMN id_nit VARCHAR(80)'))

    config_columns = {
        row[1]
        for row in db.session.execute(db.text('PRAGMA table_info(config)')).fetchall()
    }
    if 'invoice_description' not in config_columns:
        db.session.execute(db.text('ALTER TABLE config ADD COLUMN invoice_description TEXT'))
    if 'invoice_template' not in config_columns:
        db.session.execute(db.text('ALTER TABLE config ADD COLUMN invoice_template TEXT'))
    if 'receipt_template' not in config_columns:
        db.session.execute(db.text('ALTER TABLE config ADD COLUMN receipt_template TEXT'))
    if 'payment_accounts' not in config_columns:
        db.session.execute(db.text('ALTER TABLE config ADD COLUMN payment_accounts TEXT'))
    db.session.commit()

    user_columns = {
        row[1]
        for row in db.session.execute(db.text('PRAGMA table_info(user)')).fetchall()
    }
    if 'role' not in user_columns:
        db.session.execute(db.text("ALTER TABLE user ADD COLUMN role VARCHAR(30) NOT NULL DEFAULT 'admin'"))
    if 'active' not in user_columns:
        db.session.execute(db.text('ALTER TABLE user ADD COLUMN active BOOLEAN NOT NULL DEFAULT 1'))
    if 'created_at' not in user_columns:
        db.session.execute(db.text('ALTER TABLE user ADD COLUMN created_at DATETIME'))

    # Índices para acelerar búsquedas/filtros frecuentes y joins de alto uso.
    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_inventory_status ON inventory(status)'))
    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_inventory_container_id ON inventory(container_id)'))
    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_inventory_model ON inventory(model)'))
    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_inventory_imei_secondary ON inventory(imei_secondary)'))

    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_product_inventory_id ON product(inventory_id)'))
    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_product_sale_id ON product(sale_id)'))

    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_sale_client_id ON sale(client_id)'))
    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_sale_status ON sale(status)'))
    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_sale_balance ON sale(balance)'))

    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_payment_sale_id ON payment(sale_id)'))

    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_inventory_movement_inventory_id ON inventory_movement(inventory_id)'))
    db.session.execute(db.text('CREATE INDEX IF NOT EXISTS idx_inventory_movement_created_at ON inventory_movement(created_at)'))
    db.session.commit()


def init_db():
    os.makedirs('data', exist_ok=True)
    # Usar el contexto de la aplicación para operaciones de base de datos
    with app.app_context():
        db.create_all()
        ensure_schema_compatibility()
        if not Item.query.first():
            db.session.add(Item(name='Elemento de ejemplo'))
        # crear usuario por defecto si no existe
        if not User.query.first():
            admin = User(username='admin', password_hash=generate_password_hash('admin'), role='admin')
            db.session.add(admin)
        # crear configuración por defecto si no existe
        if not Config.query.first():
            cfg = Config(
                company_name='Mi Empresa',
                company_phone='',
                company_email='',
                company_address='',
                invoice_description='',
                invoice_template='',
                receipt_template='',
                payment_accounts=''
            )
            db.session.add(cfg)

        default_container = InventoryContainer.query.filter_by(name='General').first()
        if not default_container:
            default_container = InventoryContainer(name='General', description='Contenedor por defecto')
            db.session.add(default_container)
            db.session.flush()

        Inventory.query.filter(Inventory.container_id.is_(None)).update(
            {'container_id': default_container.id}, synchronize_session=False
        )
        db.session.commit()


@app.route('/')
def index():
    # Requiere login
    if not session.get('user'):
        return redirect(url_for('login'))
    # dashboard links
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            session['user'] = user.username
            if not user.active:
                session.pop('user', None)
                flash('Tu cuenta está desactivada. Contacta al administrador.', 'error')
                return render_template('login.html')
            log_action('login', detail=f'Inicio de sesión exitoso')
            db.session.commit()
            flash('Login exitoso.', 'success')
            return redirect(url_for('index'))
        flash('Usuario o contraseña incorrectos.', 'error')
    return render_template('login.html')


@app.route('/logout')
def logout():
    log_action('logout', detail='Cierre de sesión')
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
    session.pop('user', None)
    return redirect(url_for('login'))


# Clientes
@app.route('/clients')
@require_perm('ver_clientes')
def clients():
    if not session.get('user'):
        return redirect(url_for('login'))
    clients = Client.query.order_by(Client.name).all()
    clients_data = []
    for client in clients:
        sales = [s for s in client.sales if s.status != 'devuelto']
        net_balance = sum([float(s.balance) for s in sales]) if sales else 0.0
        if net_balance > 0:
            status = 'debe'
        elif net_balance < 0:
            status = 'saldo a favor'
        else:
            status = 'al dia'
        clients_data.append({
            'client': client,
            'status': status,
            'net_balance': net_balance
        })
    return render_template('clients.html', clients=clients_data)


@app.route('/clients/new', methods=['GET', 'POST'])
@require_perm('crear_clientes')
def new_client():
    if not session.get('user'):
        return redirect(url_for('login'))
    if request.method == 'POST':
        c = Client(
            name=request.form.get('name'),
            id_nit=request.form.get('id_nit'),
            phone=request.form.get('phone'),
            email=request.form.get('email'),
            address=request.form.get('address'),
            notes=request.form.get('notes')
        )
        db.session.add(c)
        db.session.commit()
        return redirect(url_for('clients'))
    return render_template('client_form.html')


@app.route('/clients/<int:client_id>')
@require_perm('ver_clientes')
def view_client(client_id):
    if not session.get('user'):
        return redirect(url_for('login'))
    client = Client.query.get_or_404(client_id)
    return render_template('client_view.html', client=client, payment_accounts=get_payment_accounts())


@app.route('/inventory', methods=['GET', 'POST'])
@require_perm('ver_inventario')
def inventory():
    if not session.get('user'):
        return redirect(url_for('login'))

    containers = InventoryContainer.query.filter_by(active=True).order_by(InventoryContainer.name.asc()).all()
    default_container = InventoryContainer.query.filter_by(name='General').first()
    if not default_container and containers:
        default_container = containers[0]

    if request.method == 'POST':
        device_type = (request.form.get('device_type') or '').strip()
        imei = (request.form.get('imei') or '').strip()
        imei_secondary = (request.form.get('imei_secondary') or '').strip()
        model = (request.form.get('model') or '').strip()
        color = (request.form.get('color') or '').strip()
        invoice_number = (request.form.get('invoice_number') or '').strip()
        container_id_raw = request.form.get('container_id')
        purchase_price = parse_currency(request.form.get('purchase_price'))
        price = parse_currency(request.form.get('price'))

        if 'ver_precio_compra' not in get_user_perms():
            purchase_price = price

        container = None
        if container_id_raw:
            try:
                container = InventoryContainer.query.filter_by(id=int(container_id_raw), active=True).first()
            except (TypeError, ValueError):
                container = None
        if not container:
            container = default_container

        if not device_type:
            flash('Selecciona el tipo de dispositivo.', 'error')
            return redirect(url_for('inventory'))
        if not imei:
            flash('El IMEI/Serie es obligatorio.', 'error')
            return redirect(url_for('inventory'))
        if device_type == 'Telefono':
            if not imei_secondary:
                flash('Para teléfonos debes ingresar IMEI 1 e IMEI 2.', 'error')
                return redirect(url_for('inventory'))
            if imei_secondary == imei:
                flash('IMEI 1 e IMEI 2 deben ser diferentes.', 'error')
                return redirect(url_for('inventory'))
        else:
            imei_secondary = None
        if purchase_price <= 0:
            flash('Ingresa un precio de compra válido mayor a 0.', 'error')
            return redirect(url_for('inventory'))
        if price <= 0:
            flash('Ingresa un precio válido mayor a 0.', 'error')
            return redirect(url_for('inventory'))
        existing_imei = Inventory.query.filter(
            db.or_(Inventory.imei == imei, Inventory.imei_secondary == imei)
        ).first()
        if existing_imei:
            if existing_imei.status == 'disponible':
                flash(f'Ese IMEI/Serie ya existe en inventario y está disponible (#{existing_imei.id} — {existing_imei.model or existing_imei.imei}).', 'error')
            else:
                flash(
                    f'Ese IMEI/Serie ya existe en inventario y fue vendido anteriormente '
                    f'(#{existing_imei.id} — {existing_imei.model or existing_imei.imei}). '
                    'No se guardó un equipo nuevo: usa la opción de reintegro para volver a ingresarlo.',
                    'error'
                )
            return redirect(url_for('inventory'))
        if imei_secondary and Inventory.query.filter(
            db.or_(Inventory.imei == imei_secondary, Inventory.imei_secondary == imei_secondary)
        ).first():
            flash('El IMEI 2 ya existe en inventario.', 'error')
            return redirect(url_for('inventory'))

        item = Inventory(
            device_type=device_type,
            imei=imei,
            imei_secondary=imei_secondary,
            model=model,
            color=color,
            invoice_number=invoice_number,
            purchase_price=purchase_price,
            price=price,
            container_id=container.id if container else None,
            status='disponible'
        )
        db.session.add(item)

        db.session.flush()
        movement = InventoryMovement(
            inventory_id=item.id,
            action='ingreso',
            from_container_id=None,
            to_container_id=item.container_id,
            note='Ingreso inicial al inventario',
            username=session.get('user')
        )
        db.session.add(movement)
        db.session.commit()
        flash('Equipo agregado a inventario.', 'success')
        log_action('agregar_inventario', resource_type='inventory', resource_id=item.id,
               detail=f'{item.device_type} | IMEI: {item.imei} | {item.model} | Precio: {item.price}')
        db.session.commit()
        return redirect(url_for('inventory'))

    search_imei = (request.args.get('search_imei') or '').strip()
    selected_container_id = (request.args.get('container_id') or '').strip()

    items_query = Inventory.query.options(joinedload(Inventory.container))
    if search_imei:
        search_like = f'%{search_imei}%'
        items_query = items_query.filter(
            db.or_(
                Inventory.imei.ilike(search_like),
                Inventory.imei_secondary.ilike(search_like),
                Inventory.model.ilike(search_like),
                Inventory.invoice_number.ilike(search_like)
            )
        )

    if selected_container_id:
        try:
            items_query = items_query.filter(Inventory.container_id == int(selected_container_id))
        except ValueError:
            selected_container_id = ''

    items = items_query.order_by(
        Inventory.status.asc(), Inventory.created_at.desc(), Inventory.id.desc()
    ).all()
    available_count = len([item for item in items if item.status == 'disponible'])
    sold_count = len([item for item in items if item.status == 'vendido'])

    assigned_clients = {}
    item_ids = [item.id for item in items]
    if item_ids:
        latest_product_per_item = db.session.query(
            Product.inventory_id.label('inventory_id'),
            db.func.max(Product.id).label('latest_product_id')
        ).join(Sale, Product.sale_id == Sale.id).filter(
            Product.inventory_id.in_(item_ids),
            Sale.status != 'devuelto'
        ).group_by(Product.inventory_id).subquery()

        assignments = db.session.query(
            Product.inventory_id,
            Client.name
        ).join(Sale, Product.sale_id == Sale.id).join(
            Client, Sale.client_id == Client.id
        ).join(
            latest_product_per_item,
            db.and_(
                Product.inventory_id == latest_product_per_item.c.inventory_id,
                Product.id == latest_product_per_item.c.latest_product_id,
            )
        ).all()

        assigned_clients = {inventory_id: client_name for inventory_id, client_name in assignments}

    invoice_info = {}
    if item_ids:
        latest_product_ids = db.session.query(
            db.func.max(Product.id).label('latest_product_id')
        ).filter(
            Product.inventory_id.in_(item_ids)
        ).group_by(Product.inventory_id).subquery()

        latest_products = Product.query.filter(
            Product.id.in_(db.session.query(latest_product_ids.c.latest_product_id))
        ).all()

        invoice_info = {
            p.inventory_id: {'sale_id': p.sale_id, 'product_id': p.id, 'paid': bool(p.paid)}
            for p in latest_products
        }

    movements_query = InventoryMovement.query.join(Inventory).order_by(
        InventoryMovement.created_at.desc(), InventoryMovement.id.desc()
    )
    if selected_container_id:
        try:
            selected_id = int(selected_container_id)
            movements_query = movements_query.filter(
                db.or_(
                    InventoryMovement.from_container_id == selected_id,
                    InventoryMovement.to_container_id == selected_id
                )
            )
        except ValueError:
            pass
    recent_movements = movements_query.limit(30).all()

    return render_template(
        'inventory.html',
        items=items,
        containers=containers,
        available_count=available_count,
        sold_count=sold_count,
        search_imei=search_imei,
        selected_container_id=selected_container_id,
        assigned_clients=assigned_clients,
        invoice_info=invoice_info,
        recent_movements=recent_movements
    )


@app.route('/inventory/check-imei')
def check_imei():
    if not session.get('user'):
        from flask import jsonify
        return jsonify({'error': 'unauthorized'}), 401
    from flask import jsonify
    imei = (request.args.get('imei') or '').strip()
    if not imei:
        return jsonify({'found': False})
    found = Inventory.query.filter(
        db.or_(Inventory.imei == imei, Inventory.imei_secondary == imei)
    ).first()
    if not found:
        return jsonify({'found': False})
    return jsonify({
        'found': True,
        'status': found.status,
        'item_id': found.id,
        'model': found.model or '',
        'imei': found.imei,
        'device_type': found.device_type or '',
        'container': found.container.name if found.container else ''
    })


@app.route('/inventory/reintegrate/<int:item_id>', methods=['POST'])
def reintegrate_inventory_item(item_id):
    if not session.get('user'):
        return redirect(url_for('login'))

    item = Inventory.query.get_or_404(item_id)
    if item.status != 'vendido':
        flash('Solo se pueden reintegrar equipos marcados como vendidos.', 'error')
        return redirect(url_for('inventory'))

    new_purchase_price = parse_currency(request.form.get('purchase_price'))
    new_price = parse_currency(request.form.get('price'))
    container_id_raw = request.form.get('container_id')
    note = (request.form.get('note') or '').strip()

    if new_purchase_price <= 0:
        flash('Ingresa un precio de compra válido para el reintegro.', 'error')
        return redirect(url_for('inventory'))
    if new_price <= 0:
        flash('Ingresa un precio de venta válido para el reintegro.', 'error')
        return redirect(url_for('inventory'))

    container = None
    if container_id_raw:
        try:
            container = InventoryContainer.query.filter_by(id=int(container_id_raw), active=True).first()
        except (TypeError, ValueError):
            container = None
    if not container:
        container = InventoryContainer.query.filter_by(name='General').first()

    item.purchase_price = new_purchase_price
    item.price = new_price
    item.status = 'disponible'
    item.sold_at = None
    if container:
        item.container_id = container.id

    db.session.add(InventoryMovement(
        inventory_id=item.id,
        action='reintegro',
        from_container_id=None,
        to_container_id=item.container_id,
        note=note or 'Reintegro al inventario',
        username=session.get('user')
    ))
    db.session.commit()
    flash(f'Equipo #{item.id} ({item.model or item.imei}) reintegrado al inventario.', 'success')
    return redirect(url_for('inventory'))


@app.route('/inventory/move/<int:item_id>', methods=['POST'])
@require_perm('mover_inventario')
def move_inventory_item(item_id):
    if not session.get('user'):
        return redirect(url_for('login'))

    item = Inventory.query.get_or_404(item_id)
    if item.status != 'disponible':
        flash('Solo se pueden trasladar equipos disponibles.', 'error')
        return redirect(url_for('inventory'))

    target_raw = request.form.get('target_container_id')
    note = (request.form.get('move_note') or '').strip()
    try:
        target_id = int(target_raw)
    except (TypeError, ValueError):
        flash('Selecciona un contenedor destino válido.', 'error')
        return redirect(url_for('inventory'))

    target = InventoryContainer.query.filter_by(id=target_id, active=True).first()
    if not target:
        flash('El contenedor destino no existe o está inactivo.', 'error')
        return redirect(url_for('inventory'))
    if item.container_id == target.id:
        flash('El equipo ya está en ese contenedor.', 'warning')
        return redirect(url_for('inventory'))

    movement = InventoryMovement(
        inventory_id=item.id,
        action='traslado',
        from_container_id=item.container_id,
        to_container_id=target.id,
        note=note or 'Traslado manual',
        username=session.get('user')
    )
    item.container_id = target.id
    db.session.add(movement)
    db.session.commit()
    flash('Equipo trasladado correctamente.', 'success')
    return redirect(url_for('inventory'))


@app.route('/inventory/move-bulk', methods=['POST'])
@require_perm('mover_inventario')
def move_inventory_items_bulk():
    if not session.get('user'):
        return redirect(url_for('login'))

    target_raw = request.form.get('target_container_id')
    note = (request.form.get('move_note') or '').strip()
    selected_raw_ids = request.form.getlist('inventory_ids[]')

    if not selected_raw_ids:
        flash('Selecciona al menos un equipo para trasladar.', 'error')
        return redirect(url_for('inventory'))

    try:
        target_id = int(target_raw)
    except (TypeError, ValueError):
        flash('Selecciona un contenedor destino válido.', 'error')
        return redirect(url_for('inventory'))

    target = InventoryContainer.query.filter_by(id=target_id, active=True).first()
    if not target:
        flash('El contenedor destino no existe o está inactivo.', 'error')
        return redirect(url_for('inventory'))

    moved = 0
    skipped = 0
    for raw_id in selected_raw_ids:
        try:
            item_id = int(raw_id)
        except (TypeError, ValueError):
            skipped += 1
            continue

        item = Inventory.query.get(item_id)
        if not item or item.status != 'disponible' or item.container_id == target.id:
            skipped += 1
            continue

        movement = InventoryMovement(
            inventory_id=item.id,
            action='traslado',
            from_container_id=item.container_id,
            to_container_id=target.id,
            note=note or 'Traslado masivo',
            username=session.get('user')
        )
        item.container_id = target.id
        db.session.add(movement)
        moved += 1

    db.session.commit()

    if moved == 0:
        flash('No se pudo trasladar ningún equipo. Revisa selección y estado.', 'warning')
    elif skipped > 0:
        flash(f'Se trasladaron {moved} equipo(s). Se omitieron {skipped}.', 'success')
    else:
        flash(f'Se trasladaron {moved} equipo(s) correctamente.', 'success')
    return redirect(url_for('inventory'))


@app.route('/inventory/containers/new', methods=['POST'])
@require_perm('mover_inventario')
def new_inventory_container():
    if not session.get('user'):
        return redirect(url_for('login'))

    name = (request.form.get('name') or '').strip()
    description = (request.form.get('description') or '').strip()
    if not name:
        flash('El nombre del contenedor es obligatorio.', 'error')
        return redirect(url_for('inventory'))

    existing = InventoryContainer.query.filter(db.func.lower(InventoryContainer.name) == name.lower()).first()
    if existing:
        flash('Ya existe un contenedor con ese nombre.', 'error')
        return redirect(url_for('inventory'))

    container = InventoryContainer(name=name, description=description or None, active=True)
    db.session.add(container)
    db.session.commit()
    flash('Contenedor creado correctamente.', 'success')
    return redirect(url_for('inventory'))


# Ventas
@app.route('/sales/new', methods=['GET', 'POST'])
@require_perm('crear_ventas')
def new_sale():
    if not session.get('user'):
        return redirect(url_for('login'))
    
    clients = Client.query.order_by(Client.name).all()
    available_inventory = Inventory.query.filter_by(status='disponible').order_by(
        Inventory.container_id.asc(), Inventory.device_type.asc(), Inventory.created_at.asc()
    ).all()
    
    client_id = request.args.get('client_id') or request.form.get('client_id')
    client = None
    if client_id:
        try:
            client = Client.query.get(int(client_id))
        except (ValueError, TypeError):
            client = None
    
    if request.method == 'POST':
        # ensure client_id is provided and valid
        try:
            cid = int(request.form.get('client_id'))
        except (TypeError, ValueError):
            flash('Debes seleccionar un cliente.', 'error')
            return render_template('sale_form.html', clients=clients, client=None)
        
        selected_inventory_ids = request.form.getlist('inventory_ids[]')
        if not selected_inventory_ids:
            flash('Debes seleccionar al menos un equipo del inventario.', 'error')
            return render_template(
                'sale_form.html',
                clients=clients,
                client=client,
                available_inventory=available_inventory
            )

        selected_items = []
        selected_ids = set()
        for raw_id in selected_inventory_ids:
            try:
                inv_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if inv_id in selected_ids:
                continue
            item = Inventory.query.filter_by(id=inv_id, status='disponible').first()
            if item:
                selected_items.append(item)
                selected_ids.add(inv_id)

        if not selected_items:
            flash('Los equipos seleccionados ya no están disponibles.', 'error')
            return render_template(
                'sale_form.html',
                clients=clients,
                client=client,
                available_inventory=available_inventory
            )

        total_price = sum([float(item.price or 0) for item in selected_items])

        initial = parse_currency(request.form.get('initial'))
        
        # Aplicar saldo a favor del cliente a la nueva venta
        credit_sales = Sale.query.filter(
            Sale.client_id == cid,
            Sale.balance < 0,
            Sale.status != 'devuelto'
        ).order_by(Sale.date.asc()).all()

        available_credit = sum([-float(s.balance) for s in credit_sales])
        due_amount = total_price - initial
        credit_applied = 0
        credit_applied_details = []
        if due_amount > 0 and available_credit > 0:
            credit_applied = min(available_credit, due_amount)
            due_amount -= credit_applied

        # Crear la venta
        sale = Sale(
            client_id=cid,
            sale_price=total_price,
            initial_amount=initial,
            balance=due_amount
        )
        if sale.balance < 0:
            sale.status = 'saldo a favor'
        elif sale.balance == 0:
            sale.status = 'pagada'
        db.session.add(sale)
        db.session.flush()

        for item in selected_items:
            source_container_id = item.container_id
            product = Product(
                sale_id=sale.id,
                inventory_id=item.id,
                imei=item.imei,
                imei_secondary=item.imei_secondary,
                device_type=item.device_type,
                color=item.color,
                invoice_number=item.invoice_number,
                model=item.model,
                purchase_price=float(item.purchase_price or 0),
                price=float(item.price or 0)
            )
            db.session.add(product)
            item.status = 'vendido'
            item.sold_at = datetime.now()
            db.session.add(InventoryMovement(
                inventory_id=item.id,
                action='venta',
                from_container_id=source_container_id,
                to_container_id=None,
                note=f'Venta #{sale.id} a {sale.client.name}',
                username=session.get('user')
            ))
        
        # Registrar pago por saldo a favor aplicado
        if credit_applied > 0:
            # Consumir el saldo a favor de ventas anteriores
            remaining_credit = credit_applied
            for credit_sale in credit_sales:
                if remaining_credit <= 0:
                    break
                sale_credit = -float(credit_sale.balance)
                consume = min(sale_credit, remaining_credit)
                credit_sale.balance = float(credit_sale.balance) + consume
                if credit_sale.balance == 0:
                    credit_sale.status = 'pagada'
                else:
                    credit_sale.status = 'saldo a favor'
                credit_applied_details.append({
                    'id': credit_sale.id,
                    'amount': consume
                })
                remaining_credit -= consume

            details_text = ''
            if credit_applied_details:
                detail_parts = [
                    f"venta #{d['id']} COP {d['amount']:,.0f}" for d in credit_applied_details
                ]
                details_text = ' | Origen: ' + '; '.join(detail_parts)

            pay_credit = Payment(
                sale_id=sale.id,
                amount=credit_applied,
                method='saldo a favor',
                note=f"Cruce de saldo a favor COP {credit_applied:,.0f}{details_text}"
            )
            db.session.add(pay_credit)

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash('No se pudo registrar la venta: uno o más equipos ya fueron vendidos por otro proceso.', 'error')
            return render_template(
                'sale_form.html',
                clients=clients,
                client=client,
                available_inventory=available_inventory
            )
        
        # si hubo pago inicial, crear payment
        if initial > 0:
            pay = Payment(sale_id=sale.id, amount=initial, method='inicial', note='Pago inicial')
            db.session.add(pay)
            db.session.commit()

        update_paid_products(sale)
        
        flash(f'Venta registrada con {len(selected_items)} dispositivo(s).', 'success')
        log_action('crear_venta', resource_type='sale', resource_id=sale.id,
               detail=f'Cliente: {client.name} | Equipos: {len(selected_items)} | Total: {sale.sale_price}')
        db.session.commit()
        return redirect(url_for('view_client', client_id=sale.client_id))

    return render_template(
        'sale_form.html',
        client=client,
        clients=clients,
        available_inventory=available_inventory
    )


@app.route('/sales/<int:sale_id>')
@require_perm('ver_ventas')
def view_sale(sale_id):
    if not session.get('user'):
        return redirect(url_for('login'))
    sale = Sale.query.get_or_404(sale_id)
    return render_template('sale_view.html', sale=sale, payment_accounts=get_payment_accounts())


# Función auxiliar para actualizar productos pagados
def update_paid_products(sale, do_commit=False):
    """Actualiza el estado 'paid' de los productos según cuánto se ha pagado de la venta"""
    total_paid = float(sale.sale_price) - float(sale.balance)
    
    # Ordenar productos por id (orden de creación)
    products = sorted(sale.products, key=lambda p: p.id)
    
    accumulated = 0.0
    for product in products:
        product_price = float(product.price)
        # Si el total pagado cubre este producto completamente
        if accumulated + product_price <= total_paid:
            product.paid = True
            accumulated += product_price
        else:
            product.paid = False
    
    if do_commit:
        db.session.commit()


# Pagos
@app.route('/payments/new', methods=['POST'])
@require_perm('registrar_pagos')
def new_payment():
    if not session.get('user'):
        return redirect(url_for('login'))
    sale_id = int(request.form.get('sale_id'))
    amount = parse_currency(request.form.get('amount'))
    method = request.form.get('method')
    account_from = (request.form.get('account_from') or '').strip()
    account_to = (request.form.get('account_to') or '').strip()
    note = request.form.get('note')
    sale = Sale.query.get_or_404(sale_id)

    if method == 'cuentas':
        if not account_from or not account_to:
            flash('Para transferencia debes seleccionar cuenta origen y cuenta destino.', 'error')
            return redirect(url_for('view_sale', sale_id=sale_id))
        if account_from == account_to:
            flash('La cuenta origen y destino deben ser diferentes.', 'error')
            return redirect(url_for('view_sale', sale_id=sale_id))

    pay = Payment(
        sale_id=sale_id,
        amount=amount,
        method=method,
        account_from=account_from or None,
        account_to=account_to or None,
        note=note,
        username=session.get('user')
    )
    sale.balance = float(sale.balance) - amount
    # Actualizar status según el balance
    if sale.balance < 0:
        extra = -float(sale.balance)
        sale.balance = 0
        sale.status = 'pagada'

        applied_details = []

        # Aplicar excedente a otras ventas pendientes del cliente
        pending_sales = Sale.query.filter(
            Sale.client_id == sale.client_id,
            Sale.id != sale.id,
            Sale.balance > 0,
            Sale.status != 'devuelto'
        ).order_by(Sale.date.asc()).all()

        remaining = extra
        for pending in pending_sales:
            if remaining <= 0:
                break
            pending_balance = float(pending.balance)
            apply_amount = min(remaining, pending_balance)
            pending.balance = pending_balance - apply_amount
            if pending.balance == 0:
                pending.status = 'pagada'
            else:
                pending.status = 'activo'
            extra_pay = Payment(
                sale_id=pending.id,
                amount=apply_amount,
                method=method or 'abono excedente',
                account_from=account_from or None,
                account_to=account_to or None,
                note=f"Abono excedente de venta #{sale.id}",
                username=session.get('user')
            )
            db.session.add(extra_pay)
            update_paid_products(pending, do_commit=False)
            remaining -= apply_amount
            applied_details.append({
                'id': pending.id,
                'amount': apply_amount
            })

        # Si queda excedente sin deudas pendientes, mantener saldo a favor
        if remaining > 0:
            sale.balance = -remaining
            sale.status = 'saldo a favor'

        if applied_details or remaining > 0:
            detail_parts = []
            if applied_details:
                detail_parts.append('Aplicado a ' + '; '.join(
                    [f"venta #{d['id']} COP {d['amount']:,.0f}" for d in applied_details]
                ))
            if remaining > 0:
                detail_parts.append(f"Saldo a favor generado COP {remaining:,.0f}")
            extra_note = ' | '.join(detail_parts)
            if pay.note:
                pay.note = f"{pay.note} | {extra_note}"
            else:
                pay.note = extra_note
    elif sale.balance == 0:
        sale.status = 'pagada'
    else:
        sale.status = 'activo'
    db.session.add(pay)
    db.session.commit()
    
    # Actualizar productos pagados
    update_paid_products(sale, do_commit=True)
    
    return redirect(url_for('view_sale', sale_id=sale_id))


# Abono automático - distribuye a ventas más antiguas
@app.route('/payments/auto', methods=['POST'])
@require_perm('registrar_pagos')
def auto_payment():
    if not session.get('user'):
        return redirect(url_for('login'))
    
    client_id = int(request.form.get('client_id'))
    amount = parse_currency(request.form.get('amount'))
    method = request.form.get('method')
    account_from = (request.form.get('account_from') or '').strip()
    account_to = (request.form.get('account_to') or '').strip()
    note = request.form.get('note')
    
    client = Client.query.get_or_404(client_id)
    
    if method == 'cuentas':
        if not account_from or not account_to:
            flash('Para transferencia debes seleccionar cuenta origen y cuenta destino.', 'error')
            return redirect(url_for('view_client', client_id=client_id))
        if account_from == account_to:
            flash('La cuenta origen y destino deben ser diferentes.', 'error')
            return redirect(url_for('view_client', client_id=client_id))

    # Obtener ventas con saldo pendiente (balance > 0), ordenadas por fecha más antigua primero
    pending_sales = Sale.query.filter(
        Sale.client_id == client_id,
        Sale.balance > 0,
        Sale.status != 'devuelto'
    ).order_by(Sale.date.asc()).all()
    
    if not pending_sales:
        flash('Este cliente no tiene ventas con saldo pendiente.', 'warning')
        return redirect(url_for('view_client', client_id=client_id))
    
    remaining_amount = amount
    sales_paid = []
    
    # Distribuir el abono a las ventas más antiguas
    for sale in pending_sales:
        if remaining_amount <= 0:
            break
        
        balance = float(sale.balance)
        
        # Determinar cuánto abonar a esta venta
        payment_amount = min(remaining_amount, balance)
        
        # Crear el pago
        pay = Payment(
            sale_id=sale.id,
            amount=payment_amount,
            method=method or 'abono automático',
            account_from=account_from or None,
            account_to=account_to or None,
            note=f"{note or 'Abono automático'} (distribuido de COP {amount:,.0f})",
            username=session.get('user')
        )
        db.session.add(pay)
        
        # Actualizar balance
        sale.balance = balance - payment_amount
        
        # Actualizar status según el balance
        if sale.balance < 0:
            sale.status = 'saldo a favor'
        elif sale.balance == 0:
            sale.status = 'pagada'
        else:
            sale.status = 'activo'
        
        remaining_amount -= payment_amount
        sales_paid.append({
            'id': sale.id,
            'amount': payment_amount,
            'balance': sale.balance
        })
        
        # Actualizar productos pagados de esta venta
        update_paid_products(sale, do_commit=False)
    
    # Si sobra dinero, aplicarlo como saldo a favor en la venta más reciente
    if remaining_amount > 0 and pending_sales:
        last_sale = pending_sales[-1]
        extra_pay = Payment(
            sale_id=last_sale.id,
            amount=remaining_amount,
            method=method or 'abono automático',
            account_from=account_from or None,
            account_to=account_to or None,
            note=f"{note or 'Abono automático - saldo a favor'} (sobrante de COP {amount:,.0f})",
            username=session.get('user')
        )
        db.session.add(extra_pay)
        last_sale.balance = float(last_sale.balance) - remaining_amount
        last_sale.status = 'saldo a favor'
        sales_paid.append({
            'id': last_sale.id,
            'amount': remaining_amount,
            'balance': last_sale.balance
        })
    
    db.session.commit()
    
    flash(f'Abono de COP {amount:,.0f} distribuido automáticamente a {len(sales_paid)} venta(s).', 'success')
    return redirect(url_for('view_client', client_id=client_id))


# Devoluciones / Returns
@app.route('/sales/<int:sale_id>/return', methods=['POST'])
@require_perm('registrar_pagos')
def process_return(sale_id):
    if not session.get('user'):
        return redirect(url_for('login'))
    sale = Sale.query.get_or_404(sale_id)
    # calcular total pagado
    total_paid = sum([float(p.amount) for p in sale.payments]) if sale.payments else 0.0
    reason = request.form.get('reason')
    # crear registro de devolución
    ret = Return(sale_id=sale.id, reason=reason, refund_amount=total_paid)
    sale.status = 'devuelto'
    sale.balance = 0
    for product in sale.products:
        if product.inventory:
            product.inventory.status = 'disponible'
            product.inventory.sold_at = None
            db.session.add(InventoryMovement(
                inventory_id=product.inventory.id,
                action='devolucion',
                from_container_id=None,
                to_container_id=product.inventory.container_id,
                note=f'Devolución de venta #{sale.id}',
                username=session.get('user')
            ))
    db.session.add(ret)
    db.session.commit()
    return redirect(url_for('view_sale', sale_id=sale_id))


# Dashboard de ventas y caja
@app.route('/sales')
@require_perm('ver_estadisticas')
def sales_dashboard():
    
    # Obtener fecha seleccionada o usar hoy
    selected_date_str = request.args.get('date')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = datetime.now().date()
    else:
        selected_date = datetime.now().date()
    
    # Filtros de búsqueda
    search_query = request.args.get('search', '').strip()
    
    today = datetime.now().date()
    selected_start = datetime.combine(selected_date, datetime.min.time())
    selected_end = datetime.combine(selected_date, datetime.max.time())
    
    # Ventas del día seleccionado
    sales_selected_query = Sale.query.filter(
        Sale.date >= selected_start,
        Sale.date <= selected_end,
        Sale.status != 'devuelto'
    )
    if search_query:
        sales_selected_query = sales_selected_query.join(Client).outerjoin(Product).filter(
            db.or_(
                Client.name.ilike(f'%{search_query}%'),
                Client.email.ilike(f'%{search_query}%'),
                Client.phone.ilike(f'%{search_query}%'),
                Product.model.ilike(f'%{search_query}%'),
                Product.invoice_number.ilike(f'%{search_query}%'),
                Product.imei.ilike(f'%{search_query}%'),
                Product.imei_secondary.ilike(f'%{search_query}%')
            )
        ).distinct()
    sales_selected = sales_selected_query.all()
    
    # Totales
    total_sales_selected = sum([float(s.sale_price) for s in sales_selected]) if sales_selected else 0.0
    devices_sold_selected = sum([len(s.products) for s in sales_selected]) if sales_selected else 0
    
    # Pagos realizados en el día seleccionado (aunque la venta sea de otro día)
    all_payments_selected = Payment.query.filter(
        Payment.date >= selected_start,
        Payment.date <= selected_end
    ).join(Sale).filter(Sale.status != 'devuelto').all()
    
    # Agrupar pagos por método
    cash_payments = [p for p in all_payments_selected if classify_income_method(p.method) == 'efectivo']
    nequi_payments = [p for p in all_payments_selected if classify_income_method(p.method) == 'nequi']
    account_payments = [p for p in all_payments_selected if classify_income_method(p.method) == 'cuentas']

    cash_income = sum([float(p.amount) for p in cash_payments])
    nequi_income = sum([float(p.amount) for p in nequi_payments])
    account_income = sum([float(p.amount) for p in account_payments])
    
    # Calcular ingresos reales del día
    actual_income = sum([float(p.amount) for p in all_payments_selected]) if all_payments_selected else 0.0
    
    # Deudores (balance > 0) - ordenados por fecha más antigua (más atrasados primero)
    debtors_query = Sale.query.filter(Sale.balance > 0, Sale.status != 'devuelto')
    
    # Aplicar filtro de búsqueda
    if search_query:
        debtors_query = debtors_query.join(Client).outerjoin(Product).filter(
            db.or_(
                Client.name.ilike(f'%{search_query}%'),
                Client.email.ilike(f'%{search_query}%'),
                Client.phone.ilike(f'%{search_query}%'),
                Product.model.ilike(f'%{search_query}%'),
                Product.invoice_number.ilike(f'%{search_query}%'),
                Product.imei.ilike(f'%{search_query}%'),
                Product.imei_secondary.ilike(f'%{search_query}%')
            )
        ).distinct()
    
    debtors = debtors_query.order_by(Sale.date.asc()).all()
    
    # Total de deuda de todos los usuarios
    total_debt_all = sum([float(s.balance) for s in debtors]) if debtors else 0.0
    
    # Alertas: ventas con más de 3 días sin ser pagadas
    three_days_ago = datetime.now() - timedelta(days=3)
    overdue = [s for s in debtors if s.date and s.date < three_days_ago]
    overdue_client_ids = {sale.client_id for sale in overdue}

    inventory_available = Inventory.query.filter_by(status='disponible').count()
    
    # Calcular días navegación
    prev_day = selected_date - timedelta(days=1)
    next_day = selected_date + timedelta(days=1)
    
    return render_template('sales_dashboard.html',
                          today=today,
                          selected_date=selected_date,
                          is_today=(selected_date == today),
                          prev_day=prev_day,
                          next_day=next_day,
                          now=datetime.now(),
                          sales_selected=sales_selected,
                          total_sales_today=total_sales_selected,
                          phones_sold_today=devices_sold_selected,
                          actual_income=actual_income,
                          cash_income=cash_income,
                          nequi_income=nequi_income,
                          account_income=account_income,
                          cash_payments=cash_payments,
                          nequi_payments=nequi_payments,
                          account_payments=account_payments,
                          total_debt_all=total_debt_all,
                          debtors=debtors,
                          overdue=overdue,
                          overdue_clients_count=len(overdue_client_ids),
                          inventory_available=inventory_available,
                          search_query=search_query)


# Crear usuario
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not username or len(username) < 3:
            flash('El usuario debe tener al menos 3 caracteres.', 'error')
            return render_template('register.html')
        
        if len(password) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'error')
            return render_template('register.html')
        
        if password != confirm_password:
            flash('Las contraseñas no coinciden.', 'error')
            return render_template('register.html')
        
        if User.query.filter_by(username=username).first():
            flash('El usuario ya existe.', 'error')
            return render_template('register.html')
        
        user = User(username=username, password_hash=generate_password_hash(password))
        db.session.add(user)
        db.session.commit()
        flash('Usuario creado. Puedes iniciar sesión ahora.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')


# Configuración
@app.route('/settings', methods=['GET', 'POST'])
@require_perm('ver_configuracion')
def settings():
    
    cfg = Config.query.first() or Config()
    
    if request.method == 'POST':
        cfg.company_name = request.form.get('company_name') or cfg.company_name
        cfg.company_phone = request.form.get('company_phone')
        cfg.company_email = request.form.get('company_email')
        cfg.company_address = request.form.get('company_address')
        cfg.invoice_description = request.form.get('invoice_description')
        cfg.invoice_template = request.form.get('invoice_template')
        cfg.receipt_template = request.form.get('receipt_template')
        posted_accounts = [
            value.strip()
            for value in request.form.getlist('payment_accounts[]')
            if (value or '').strip()
        ]
        if not posted_accounts:
            # Compatibilidad con formulario anterior (textarea única)
            raw_accounts = request.form.get('payment_accounts') or ''
            posted_accounts = [line.strip() for line in raw_accounts.splitlines() if line.strip()]
        cfg.payment_accounts = '\n'.join(posted_accounts)
        
        # Manejar carga de logo
        if 'logo' in request.files:
            file = request.files['logo']
            if file and file.filename and '.' in file.filename:
                filename = secure_filename(file.filename)
                os.makedirs(UPLOADS_DIR, exist_ok=True)
                filepath = os.path.join(UPLOADS_DIR, filename)
                file.save(filepath)
                cfg.logo_path = f'uploads/{filename}'
        
        db.session.add(cfg)
        db.session.commit()
        flash('Configuración actualizada.', 'success')
        return redirect(url_for('settings'))
    
    payment_accounts_list = [line.strip() for line in (cfg.payment_accounts or '').splitlines() if line.strip()]
    return render_template('settings.html', config=cfg, payment_accounts_list=payment_accounts_list)


# Exportar Excel
# ---- Gestión de usuarios ----
@app.route('/settings/users')
@require_perm('gestionar_usuarios')
def manage_users():
    users = User.query.order_by(User.username).all()
    return render_template('users.html', users=users, role_labels=ROLE_LABELS, roles=list(ROLE_PERMISSIONS.keys()))


@app.route('/settings/users/new', methods=['POST'])
@require_perm('gestionar_usuarios')
def create_user():
    username = (request.form.get('username') or '').strip().lower()
    password = request.form.get('password') or ''
    role = request.form.get('role') or 'vendedor'
    if not username or not password:
        flash('Usuario y contraseña son obligatorios.', 'error')
        return redirect(url_for('manage_users'))
    if role not in ROLE_PERMISSIONS:
        flash('Rol inválido.', 'error')
        return redirect(url_for('manage_users'))
    if len(password) < 6:
        flash('La contraseña debe tener al menos 6 caracteres.', 'error')
        return redirect(url_for('manage_users'))
    if User.query.filter_by(username=username).first():
        flash(f'El usuario "{username}" ya existe.', 'error')
        return redirect(url_for('manage_users'))
    user = User(username=username, password_hash=generate_password_hash(password), role=role, active=True)
    db.session.add(user)
    log_action('crear_usuario', resource_type='user', detail=f'Usuario: {username}, Rol: {role}')
    db.session.commit()
    flash(f'Usuario "{username}" creado con rol {ROLE_LABELS.get(role, role)}.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/settings/users/<int:user_id>/edit', methods=['POST'])
@require_perm('gestionar_usuarios')
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    new_role = request.form.get('role') or user.role
    new_password = (request.form.get('password') or '').strip()
    new_active = request.form.get('active') == '1'
    if user.role == 'admin' and new_role != 'admin':
        admin_count = User.query.filter_by(role='admin', active=True).count()
        if admin_count <= 1:
            flash('Debe existir al menos un administrador activo.', 'error')
            return redirect(url_for('manage_users'))
    if user.role == 'admin' and not new_active:
        admin_count = User.query.filter_by(role='admin', active=True).count()
        if admin_count <= 1:
            flash('No puedes desactivar al único administrador.', 'error')
            return redirect(url_for('manage_users'))
    changes = []
    if new_role != user.role:
        changes.append(f'rol: {user.role}→{new_role}')
        user.role = new_role
    if new_active != user.active:
        changes.append(f'activo: {user.active}→{new_active}')
        user.active = new_active
    if new_password:
        if len(new_password) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'error')
            return redirect(url_for('manage_users'))
        user.password_hash = generate_password_hash(new_password)
        changes.append('contraseña actualizada')
    if changes:
        log_action('editar_usuario', resource_type='user', resource_id=user_id,
                   detail=f'Usuario: {user.username} | Cambios: {", ".join(changes)}')
        db.session.commit()
        flash(f'Usuario "{user.username}" actualizado.', 'success')
    else:
        flash('Sin cambios.', '')
    return redirect(url_for('manage_users'))


@app.route('/settings/users/<int:user_id>/delete', methods=['POST'])
@require_perm('gestionar_usuarios')
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.username == session.get('user'):
        flash('No puedes eliminar tu propia cuenta.', 'error')
        return redirect(url_for('manage_users'))
    if user.role == 'admin':
        admin_count = User.query.filter_by(role='admin', active=True).count()
        if admin_count <= 1:
            flash('No puedes eliminar al único administrador.', 'error')
            return redirect(url_for('manage_users'))
    log_action('eliminar_usuario', resource_type='user', resource_id=user_id,
               detail=f'Usuario eliminado: {user.username} (rol: {user.role})')
    db.session.delete(user)
    db.session.commit()
    flash(f'Usuario "{user.username}" eliminado.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/audit-log')
@require_perm('ver_auditoria')
def audit_log():
    page = int(request.args.get('page', 1))
    per_page = 50
    filter_user = request.args.get('filter_user', '').strip()
    filter_action = request.args.get('filter_action', '').strip()
    q = AuditLog.query.order_by(AuditLog.timestamp.desc())
    if filter_user:
        q = q.filter(AuditLog.username == filter_user)
    if filter_action:
        q = q.filter(AuditLog.action.ilike(f'%{filter_action}%'))
    total = q.count()
    logs = q.offset((page - 1) * per_page).limit(per_page).all()
    users_list = [u.username for u in User.query.order_by(User.username).all()]
    pages = (total + per_page - 1) // per_page
    return render_template('audit_log.html', logs=logs, page=page, pages=pages,
                           total=total, users_list=users_list,
                           filter_user=filter_user, filter_action=filter_action)


# Exportar Excel
@app.route('/export-excel')
@require_perm('exportar_excel')
def export_excel():
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    
    # Encabezado
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    headers = ['ID', 'Nombre', 'ID/NIT', 'Telefono', 'Email', 'Direccion', 'Notas']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Datos de clientes
    clients = Client.query.all()
    for row, client in enumerate(clients, 2):
        ws.cell(row=row, column=1, value=client.id)
        ws.cell(row=row, column=2, value=client.name)
        ws.cell(row=row, column=3, value=client.id_nit)
        ws.cell(row=row, column=4, value=client.phone)
        ws.cell(row=row, column=5, value=client.email)
        ws.cell(row=row, column=6, value=client.address)
        ws.cell(row=row, column=7, value=client.notes)
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    
    # Crear segunda hoja para ventas
    ws_sales = wb.create_sheet('Ventas')
    headers_sales = [
        'ID', 'Cliente', 'IMEI/Serial', 'Precio venta', 'Costo compra',
        'Ganancia bruta', 'Saldo', 'Estado', 'Fecha'
    ]
    for col, header in enumerate(headers_sales, 1):
        cell = ws_sales.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    sales = Sale.query.all()
    for row, sale in enumerate(sales, 2):
        total_purchase_cost = sum([float(p.purchase_price or 0) for p in sale.products]) if sale.products else 0.0
        gross_profit = float(sale.sale_price or 0) - total_purchase_cost

        ws_sales.cell(row=row, column=1, value=sale.id)
        ws_sales.cell(row=row, column=2, value=sale.client.name)
        # Concatenar IMEIs con estado de pago
        if sale.products:
            imeis = ', '.join([
                f"{'✓' if p.paid else '○'} {p.imei}{(' / ' + p.imei_secondary) if p.imei_secondary else ''}"
                for p in sale.products
            ])
        else:
            imeis = ''
        ws_sales.cell(row=row, column=3, value=imeis)
        # Precio y saldo: formato de número con separadores de miles
        price_cell = ws_sales.cell(row=row, column=4, value=int(sale.sale_price))
        price_cell.number_format = '#,##0'
        purchase_cell = ws_sales.cell(row=row, column=5, value=int(total_purchase_cost))
        purchase_cell.number_format = '#,##0'
        profit_cell = ws_sales.cell(row=row, column=6, value=int(gross_profit))
        profit_cell.number_format = '#,##0'
        balance_cell = ws_sales.cell(row=row, column=7, value=int(sale.balance))
        balance_cell.number_format = '#,##0'
        ws_sales.cell(row=row, column=8, value=sale.status)
        ws_sales.cell(row=row, column=9, value=sale.date.strftime('%d/%m/%Y') if sale.date else '')
    
    for col in range(1, len(headers_sales) + 1):
        ws_sales.column_dimensions[get_column_letter(col)].width = 18

    # Hoja de inventario completa
    ws_inventory = wb.create_sheet('Inventario')
    headers_inventory = [
        'ID', 'Contenedor', 'Tipo', 'IMEI/Serial', 'IMEI 2', 'Modelo', 'Color', 'No factura',
        'Precio compra', 'Precio venta', 'Estado', 'Cliente asignado', 'Venta ID', 'Fecha ingreso', 'Fecha vendido'
    ]
    for col, header in enumerate(headers_inventory, 1):
        cell = ws_inventory.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    inventory_items = Inventory.query.order_by(Inventory.id.asc()).all()
    for row, item in enumerate(inventory_items, 2):
        assigned_client = ''
        assigned_sale_id = ''
        related_sale = None
        for product in sorted(item.products, key=lambda p: p.id, reverse=True):
            if product.sale and product.sale.status != 'devuelto':
                related_sale = product.sale
                break
        if related_sale:
            assigned_client = related_sale.client.name if related_sale.client else ''
            assigned_sale_id = related_sale.id

        ws_inventory.cell(row=row, column=1, value=item.id)
        ws_inventory.cell(row=row, column=2, value=item.container.name if item.container else '')
        ws_inventory.cell(row=row, column=3, value=item.device_type)
        ws_inventory.cell(row=row, column=4, value=item.imei)
        ws_inventory.cell(row=row, column=5, value=item.imei_secondary)
        ws_inventory.cell(row=row, column=6, value=item.model)
        ws_inventory.cell(row=row, column=7, value=item.color)
        ws_inventory.cell(row=row, column=8, value=item.invoice_number)
        c_purchase = ws_inventory.cell(row=row, column=9, value=int(float(item.purchase_price or 0)))
        c_purchase.number_format = '#,##0'
        c_sale = ws_inventory.cell(row=row, column=10, value=int(float(item.price or 0)))
        c_sale.number_format = '#,##0'
        ws_inventory.cell(row=row, column=11, value=item.status)
        ws_inventory.cell(row=row, column=12, value=assigned_client)
        ws_inventory.cell(row=row, column=13, value=assigned_sale_id)
        ws_inventory.cell(row=row, column=14, value=item.created_at.strftime('%d/%m/%Y %H:%M') if item.created_at else '')
        ws_inventory.cell(row=row, column=15, value=item.sold_at.strftime('%d/%m/%Y %H:%M') if item.sold_at else '')

    for col in range(1, len(headers_inventory) + 1):
        ws_inventory.column_dimensions[get_column_letter(col)].width = 18

    # Hoja de pagos
    ws_payments = wb.create_sheet('Pagos')
    headers_payments = ['ID pago', 'Venta ID', 'Cliente', 'Monto', 'Metodo', 'Cuenta origen', 'Cuenta destino', 'Nota', 'Fecha']
    for col, header in enumerate(headers_payments, 1):
        cell = ws_payments.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    payments = Payment.query.order_by(Payment.date.asc()).all()
    for row, payment in enumerate(payments, 2):
        ws_payments.cell(row=row, column=1, value=payment.id)
        ws_payments.cell(row=row, column=2, value=payment.sale_id)
        ws_payments.cell(row=row, column=3, value=payment.sale.client.name if payment.sale and payment.sale.client else '')
        amount_cell = ws_payments.cell(row=row, column=4, value=int(float(payment.amount or 0)))
        amount_cell.number_format = '#,##0'
        ws_payments.cell(row=row, column=5, value=payment.method)
        ws_payments.cell(row=row, column=6, value=payment.account_from)
        ws_payments.cell(row=row, column=7, value=payment.account_to)
        ws_payments.cell(row=row, column=8, value=payment.note)
        ws_payments.cell(row=row, column=9, value=payment.date.strftime('%d/%m/%Y %H:%M') if payment.date else '')

    for col in range(1, len(headers_payments) + 1):
        ws_payments.column_dimensions[get_column_letter(col)].width = 22

    # Hoja de devoluciones
    ws_returns = wb.create_sheet('Devoluciones')
    headers_returns = ['ID devolucion', 'Venta ID', 'Cliente', 'Motivo', 'Reembolso', 'Fecha']
    for col, header in enumerate(headers_returns, 1):
        cell = ws_returns.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    returns = Return.query.order_by(Return.date.asc()).all()
    for row, item in enumerate(returns, 2):
        ws_returns.cell(row=row, column=1, value=item.id)
        ws_returns.cell(row=row, column=2, value=item.sale_id)
        ws_returns.cell(row=row, column=3, value=item.sale.client.name if item.sale and item.sale.client else '')
        ws_returns.cell(row=row, column=4, value=item.reason)
        refund_cell = ws_returns.cell(row=row, column=5, value=int(float(item.refund_amount or 0)))
        refund_cell.number_format = '#,##0'
        ws_returns.cell(row=row, column=6, value=item.date.strftime('%d/%m/%Y %H:%M') if item.date else '')

    for col in range(1, len(headers_returns) + 1):
        ws_returns.column_dimensions[get_column_letter(col)].width = 22

    # Hoja de movimientos de inventario
    ws_movements = wb.create_sheet('Movimientos inventario')
    headers_movements = [
        'ID movimiento', 'Fecha', 'Usuario', 'Accion', 'Equipo ID', 'IMEI/Serial',
        'Modelo', 'Contenedor origen', 'Contenedor destino', 'Nota'
    ]
    for col, header in enumerate(headers_movements, 1):
        cell = ws_movements.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    movements = InventoryMovement.query.order_by(InventoryMovement.created_at.asc(), InventoryMovement.id.asc()).all()
    for row, movement in enumerate(movements, 2):
        ws_movements.cell(row=row, column=1, value=movement.id)
        ws_movements.cell(row=row, column=2, value=movement.created_at.strftime('%d/%m/%Y %H:%M') if movement.created_at else '')
        ws_movements.cell(row=row, column=3, value=movement.username)
        ws_movements.cell(row=row, column=4, value=movement.action)
        ws_movements.cell(row=row, column=5, value=movement.inventory_id)
        ws_movements.cell(row=row, column=6, value=movement.inventory.imei if movement.inventory else '')
        ws_movements.cell(row=row, column=7, value=movement.inventory.model if movement.inventory else '')
        ws_movements.cell(row=row, column=8, value=movement.from_container.name if movement.from_container else '')
        ws_movements.cell(row=row, column=9, value=movement.to_container.name if movement.to_container else '')
        ws_movements.cell(row=row, column=10, value=movement.note)

    for col in range(1, len(headers_movements) + 1):
        ws_movements.column_dimensions[get_column_letter(col)].width = 22
    
    # Enviar archivo
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'reporte_completo_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# PDF de factura
@app.route('/invoice-pdf/<int:sale_id>')
def invoice_pdf(sale_id):
    if not session.get('user'):
        return redirect(url_for('login'))
    
    sale = Sale.query.get_or_404(sale_id)
    paid_products = [p for p in sale.products if p.paid]
    if len(paid_products) != 1:
        flash('Esta venta tiene varios dispositivos pagados. Genera la factura desde cada dispositivo pagado.', 'error')
        return redirect(url_for('view_sale', sale_id=sale_id))
    return generate_device_invoice_pdf(sale, paid_products[0])


@app.route('/invoice-pdf/<int:sale_id>/product/<int:product_id>')
def invoice_pdf_product(sale_id, product_id):
    if not session.get('user'):
        return redirect(url_for('login'))

    sale = Sale.query.get_or_404(sale_id)
    product = Product.query.get_or_404(product_id)
    if product.sale_id != sale.id:
        flash('El dispositivo no corresponde a esta venta.', 'error')
        return redirect(url_for('view_sale', sale_id=sale_id))
    if not product.paid:
        flash('Solo se puede imprimir factura de dispositivos cancelados.', 'error')
        return redirect(url_for('view_sale', sale_id=sale_id))

    return generate_device_invoice_pdf(sale, product)


@app.route('/payments/<int:payment_id>/receipt-pdf')
def payment_receipt_pdf(payment_id):
    if not session.get('user'):
        return redirect(url_for('login'))

    payment = Payment.query.get_or_404(payment_id)
    sale = payment.sale
    return generate_payment_receipt_pdf(sale, payment)


def generate_payment_receipt_pdf(sale, payment):
    cfg = Config.query.first()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*cm, bottomMargin=0.5*cm)
    story = []
    styles = getSampleStyleSheet()

    if cfg:
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#0f172a'), spaceAfter=4)
        company_name = cfg.company_name or 'Mi Local'
        story.append(Paragraph(f"<b>{company_name}</b>", title_style))
        info_text = f"{cfg.company_phone or '—'} • {cfg.company_email or '—'}<br/>{cfg.company_address or '—'}"
        story.append(Paragraph(info_text, styles['Normal']))

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("<b>RECIBO DE ABONO</b>", styles['Heading2']))
    story.append(Spacer(1, 0.3*cm))

    receipt_data = [
        ['Recibo #', str(payment.id), 'Fecha:', payment.date.strftime('%d/%m/%Y %H:%M') if payment.date else ''],
        ['Venta #', str(sale.id), 'Cliente:', sale.client.name],
        ['ID/NIT:', sale.client.id_nit or '—', 'Tel:', sale.client.phone or '—'],
        ['Email:', sale.client.email or '—', 'Dirección:', sale.client.address or '—'],
        ['Atendido por:', payment.username or '—', '', ''],
    ]
    receipt_table = Table(receipt_data, colWidths=[2*cm, 4*cm, 2*cm, 4*cm])
    receipt_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(receipt_table)
    story.append(Spacer(1, 0.3*cm))

    # Dispositivo(s) asociados a la venta abonada
    devices_list = []
    if sale.products:
        story.append(Paragraph("<b>Dispositivo(s) asociado(s):</b>", styles['Heading3']))
        device_rows = [['Equipo', 'IMEI/Serial', 'Bloque / Local', 'Estado']]
        for product in sale.products:
            block_name = '—'
            if product.inventory and product.inventory.container:
                block_name = product.inventory.container.name
            imei_text = product.imei or '—'
            if product.imei_secondary:
                imei_text += f" / {product.imei_secondary}"
            estado = 'Pagado' if product.paid else 'Pendiente'
            device_rows.append([product.model or product.device_type or '—', imei_text, block_name, estado])
            devices_list.append(f"{product.model or product.device_type or 'Equipo'} (IMEI {imei_text}) - Bloque: {block_name}")

        device_table = Table(device_rows, colWidths=[3.5*cm, 4.5*cm, 3*cm, 2*cm])
        device_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8f5e9')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(device_table)
        story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("<b>Detalle del abono:</b>", styles['Heading3']))
    account_detail = '—'
    if payment.method == 'cuentas' and (payment.account_from or payment.account_to):
        account_detail = f"{payment.account_from or '—'} → {payment.account_to or '—'}"

    template_text = (cfg.receipt_template or '').strip() if cfg else ''
    if template_text:
        context = {
            'empresa': (cfg.company_name or 'Mi Local') if cfg else 'Mi Local',
            'recibo_id': str(payment.id),
            'venta_id': str(sale.id),
            'fecha': payment.date.strftime('%d/%m/%Y %H:%M') if payment.date else '',
            'cliente': sale.client.name or '',
            'cliente_id_nit': sale.client.id_nit or '',
            'cliente_telefono': sale.client.phone or '',
            'cliente_email': sale.client.email or '',
            'cliente_direccion': sale.client.address or '',
            'atendido_por': payment.username or '—',
            'dispositivos': '; '.join(devices_list) if devices_list else '—',
            'monto': f"COP {float(payment.amount):,.0f}",
            'metodo': payment.method or '—',
            'cuentas': account_detail,
            'nota': payment.note or '—',
            'saldo': f"COP {float(sale.balance):,.0f}" if sale.balance >= 0 else f"Saldo a favor COP {abs(float(sale.balance)):,.0f}",
        }

        rendered = template_text
        for key, value in context.items():
            rendered = rendered.replace(f'{{{key}}}', str(value))

        story.append(Paragraph('<b>Formato de recibo:</b>', styles['Heading3']))
        for line in rendered.splitlines():
            clean_line = escape(line).replace('  ', '&nbsp;&nbsp;')
            story.append(Paragraph(clean_line if clean_line else '&nbsp;', styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
    else:
        detail_data = [
            ['Concepto', 'Detalle'],
            ['Monto abonado', f"COP {float(payment.amount):,.0f}"],
            ['Método de pago', payment.method or '—'],
            ['Cuentas', account_detail],
            ['Nota', payment.note or '—'],
            ['Saldo restante de la venta', f"COP {float(sale.balance):,.0f}" if sale.balance >= 0 else f"Saldo a favor COP {abs(float(sale.balance)):,.0f}"],
        ]
        detail_table = Table(detail_data, colWidths=[5*cm, 6*cm])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(detail_table)
        story.append(Spacer(1, 0.5*cm))

    if cfg and cfg.invoice_description:
        story.append(Paragraph("<b>Descripcion / garantia:</b>", styles['Heading3']))
        story.append(Paragraph(cfg.invoice_description.replace('\n', '<br/>'), styles['Normal']))
        story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    buffer.seek(0)

    safe_name = (sale.client.name or 'cliente').strip().replace(' ', '_')
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'recibo_abono_{safe_name}_pago_{payment.id}.pdf'
    )


def generate_device_invoice_pdf(sale, product):
    cfg = Config.query.first()
    
    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*cm, bottomMargin=0.5*cm)
    story = []
    styles = getSampleStyleSheet()
    
    # Título y datos empresa
    if cfg:
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#0f172a'), spaceAfter=4)
        company_name = cfg.company_name or 'Mi Local'
        story.append(Paragraph(f"<b>{company_name}</b>", title_style))
        info_text = f"{cfg.company_phone or '—'} • {cfg.company_email or '—'}<br/>{cfg.company_address or '—'}"
        story.append(Paragraph(info_text, styles['Normal']))
    
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("<b>FACTURA DE VENTA - DISPOSITIVO CANCELADO</b>", styles['Heading2']))
    story.append(Spacer(1, 0.3*cm))
    
    # Datos de factura
    invoice_data = [
        ['Venta #', str(sale.id), 'Fecha:', sale.date.strftime('%d/%m/%Y %H:%M') if sale.date else ''],
        ['Cliente:', sale.client.name, 'Tel:', sale.client.phone or '—'],
        ['ID/NIT:', sale.client.id_nit or '—', 'Email:', sale.client.email or '—'],
        ['Dirección:', sale.client.address or '—', '', ''],
    ]
    invoice_table = Table(invoice_data, colWidths=[2*cm, 4*cm, 2*cm, 4*cm])
    invoice_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(invoice_table)
    story.append(Spacer(1, 0.3*cm))
    
    # Datos del dispositivo cancelado
    story.append(Paragraph("<b>Dispositivo cancelado:</b>", styles['Heading3']))
    product_data = [
        ['Estado:', '✓ PAGADO', 'Tipo:', product.device_type or '—'],
        ['IMEI/Serial:', f"{product.imei}{(' / ' + product.imei_secondary) if product.imei_secondary else ''}", 'Modelo:', product.model or '—'],
        ['Color:', product.color or '—', 'Factura:', product.invoice_number or '—'],
        ['Precio:', f"COP {float(product.price):,.0f}", '', ''],
    ]
    product_table = Table(product_data, colWidths=[2*cm, 4*cm, 2*cm, 4*cm])
    product_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (3, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('BACKGROUND', (0, 0), (3, 0), colors.HexColor('#e8f5e9')),
    ]))
    story.append(product_table)
    story.append(Spacer(1, 0.1*cm))

    template_text = (cfg.invoice_template or '').strip() if cfg else ''
    if template_text:
        context = {
            'empresa': cfg.company_name or 'Mi Local' if cfg else 'Mi Local',
            'venta_id': str(sale.id),
            'fecha': sale.date.strftime('%d/%m/%Y %H:%M') if sale.date else '',
            'cliente': sale.client.name or '',
            'cliente_id_nit': sale.client.id_nit or '',
            'cliente_telefono': sale.client.phone or '',
            'cliente_email': sale.client.email or '',
            'cliente_direccion': sale.client.address or '',
            'tipo': product.device_type or '',
            'imei': product.imei or '',
            'imei2': product.imei_secondary or '',
            'modelo': product.model or '',
            'color': product.color or '',
            'factura': product.invoice_number or '',
            'precio': f"COP {float(product.price):,.0f}",
            'estado': 'CANCELADO',
        }

        rendered = template_text
        for key, value in context.items():
            rendered = rendered.replace(f'{{{key}}}', str(value))

        story.append(Paragraph('<b>Formato de factura:</b>', styles['Heading3']))
        for line in rendered.splitlines():
            clean_line = escape(line).replace('  ', '&nbsp;&nbsp;')
            story.append(Paragraph(clean_line if clean_line else '&nbsp;', styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
    else:
        # Resumen financiero por defecto
        summary_data = [
            ['Concepto', 'Monto'],
            ['Precio del dispositivo', f"COP {float(product.price):,.0f}"],
            ['Estado del dispositivo', 'CANCELADO'],
            ['Cliente', sale.client.name],
        ]
        summary_table = Table(summary_data, colWidths=[6*cm, 3*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.5*cm))

    if cfg and cfg.invoice_description:
        story.append(Paragraph("<b>Descripcion / garantia:</b>", styles['Heading3']))
        story.append(Paragraph(cfg.invoice_description.replace('\n', '<br/>'), styles['Normal']))
        story.append(Spacer(1, 0.3*cm))
    
    story.append(Paragraph("Documento generado automaticamente por el sistema.", styles['Normal']))
    
    # Compilar PDF
    doc.build(story)
    buffer.seek(0)
    
    safe_name = (sale.client.name or 'cliente').strip().replace(' ', '_')
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'factura_{safe_name}_disp_{product.id}.pdf'
    )


def open_browser():
    webbrowser.open('http://127.0.0.1:5001')


def start_backup_thread(interval_seconds=1800):
    def backup_loop():
        backup_path = os.path.join(DATA_DIR, 'app.db.backup')
        temp_path = backup_path + '.tmp'
        while True:
            try:
                os.makedirs(DATA_DIR, exist_ok=True)
                if os.path.exists(DB_PATH):
                    shutil.copy2(DB_PATH, temp_path)
                    os.replace(temp_path, backup_path)
            except Exception:
                pass
            time.sleep(interval_seconds)

    thread = threading.Thread(target=backup_loop, daemon=True)
    thread.start()


if __name__ == '__main__':
    init_db()
    # abrir navegador tras 1s para que el servidor arranque
    threading.Timer(1.0, open_browser).start()
    start_backup_thread(1800)
    app.run(host='127.0.0.1', port=5001)
